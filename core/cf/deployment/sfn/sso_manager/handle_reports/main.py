import logging
import utils
from os import getenv
from messenger import Messenger
import xlsxwriter
import openpyxl

logger = logging.getLogger()
logger.setLevel(logging.INFO)

PROJECT_NAME = getenv('PROJECT_NAME')
SSO_CROSS_ACCOUNT_ROLE_ARN = getenv('SSO_CROSS_ACCOUNT_ROLE_ARN')
SSO_USER_ID_NAMES_SECRET = getenv('SSO_USER_ID_NAMES_SECRET')
BUCKET_NAME = getenv('ARTIFACTS_BUCKET_NAME')
SENDER_EMAIL_ADDRESS = getenv('SENDER_EMAIL_ADDRESS')
RECEIVER_EMAIL_ADDRESSES = getenv('RECEIVER_EMAIL_ADDRESSES').replace(' ','').split(',')
NOTIFICATION_CONFIGS_SECRET_NAME = getenv('NOTIFICATION_CONFIGS_SECRET_NAME')
NOTIFICATION_CONFIGS = utils.get_secret_value(NOTIFICATION_CONFIGS_SECRET_NAME)
NOTIFICATION_APP = NOTIFICATION_CONFIGS.get('NOTIFICATION_APP', '')
WEBHOOK_URLS = ""
if "APP_CONFIG" in NOTIFICATION_CONFIGS:
    WEBHOOK_URLS = NOTIFICATION_CONFIGS.get("APP_CONFIG")
else:
    WEBHOOK_URLS = NOTIFICATION_CONFIGS.get("WEBHOOK_URL")
WEBHOOK_URLS = WEBHOOK_URLS.replace(' ', '').split(',')
FORMULA_PREFIXES = ('=', '+', '-', '@')

def sanitize_spreadsheet_value(value):
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value

def lambda_handler(event, context):
    logger.info('PROCESSING NEW EVENT: %s', event)

    event_name = event[0]['eventName']
    process_id = event[0]['processId']
    user_id = event[0]['userId']
    username = event[0]['userName']
    main_filename = f'/tmp/{user_id}.xlsx'
    email_sent = False
    messenger = Messenger(NOTIFICATION_APP, WEBHOOK_URLS)

    downloaded_s3_files, keys = utils.get_s3_files(BUCKET_NAME, process_id, user_id)
    accounts_done = []
    if downloaded_s3_files:
        if len(keys) > 0:
            print(keys)
            workbook_write = xlsxwriter.Workbook(main_filename)
            for key in keys:
                file_path = key.replace('.xlsx', '').split('/')
                try:
                    region = file_path[-1]
                    account = file_path[-2]
                    workbook = openpyxl.load_workbook(f'/tmp/{process_id}/{user_id}/{account}/{region}.xlsx')
                    worksheet_write = ''
                    if account not in accounts_done:
                        accounts_done.append(account)
                        worksheet_write = workbook_write.add_worksheet(f'{account}')
                    else:
                        worksheet_write = workbook_write.get_worksheet_by_name(account)
                    worksheet = workbook[account]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            row_index = cell.row - 1
                            col_index = cell.column - 1
                            if row_index == 0:
                                bold_format = workbook_write.add_format({'bold': True})
                                worksheet_write.write(row_index, col_index, sanitize_spreadsheet_value(cell.value), bold_format)
                            else:
                                worksheet_write.write(row_index, col_index, sanitize_spreadsheet_value(cell.value))
                except FileNotFoundError:
                    logger.error("No report for %s", account)
            workbook_write.close()
            if not utils.put_object(main_filename, BUCKET_NAME, f'{process_id}/{user_id}.xlsx'):
                logger.error("Could not store final report to %s S3 Bucket", BUCKET_NAME)
            for email in RECEIVER_EMAIL_ADDRESSES:
                if not messenger.send_email(SENDER_EMAIL_ADDRESS, email, username, 'disabled' if event_name == 'DisableUser' else 'deleted', main_filename):
                    logger.error("Could not send report to %s email address", email)
            email_sent = True
    else:
        logger.error("Could not get report files from %s S3 Bucket", BUCKET_NAME)
    all_sso_users = utils.get_all_sso_users_id_names(PROJECT_NAME, SSO_CROSS_ACCOUNT_ROLE_ARN)
    if not utils.store_all_sso_users_id_names(SSO_USER_ID_NAMES_SECRET, all_sso_users):
        logger.error("Could not update data for all SSO Users in SecretsManager Secret")
    if not messenger.send_alert(username, 'disabled' if event_name == 'DisableUser' else 'deleted', email_sent):
        logger.error("Could not send alert to Notification Channel")
